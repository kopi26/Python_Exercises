import sys
from openpyxl import *
from parse_cupboard_parts import *
from write_orderparts_details import *

workbook = None
cupboard_parts = {}
cupboard_orders = []





#get excel access
def access_excel(path):
    xlsx_file = path
    wb = load_workbook(xlsx_file)
    return wb



#set the data as same format and cleanup unnecessary characters
def cleanup_data(measure):
    if measure:
        measure = measure.replace('-','')
        measure = measure.replace('x',' X ')
        measure = measure.replace('X',' X ')
        #set single space    
        measure_str = ' '.join(measure.split())
        
        return measure_str
    
    

#Extract cupboard parts default measures data           
def extract_VANITY_INFO():
    ws = workbook["VANITY INFO"]

    for row in range(1, ws.max_row+1):
        code = ws.cell(row,3).value
        if code != None:
            if(isinstance(code,int)):
                measure = ws.cell(row,4).value
                measure_part = cleanup_data(measure)
                cupboard_parts[code] = split_parts_details(measure_part,code,row)


    
#get orders details
def get_Orders_MAIN_SHEET():
    ws = workbook["Main Sheet"]
    order_list = []
    
    for row in range(1, ws.max_row+1):
        order_code = ws.cell(row,5).value

        if(isinstance(order_code,int)):
            material = ws.cell(row,8).value.strip()
            style = ws.cell(row,7).value.strip()
            color = ws.cell(row,9).value.strip()
            
            if order_code in cupboard_parts:
                order_list = cupboard_parts[order_code]
                for order in order_list:
                    # Create a new order list
                    order = order[:]

                    order.insert(1,material)
                    order.insert(2,style)
                    order.insert(3,color)
                    
                    #join types by underscore
                    order[1] = '_'.join(order[1:])

                    #remove same items after join from list
                    del order[2:]

                    for parts in cupboard_orders:
                        if order[1] == parts[1]:
                            index = cupboard_orders.index(parts)
                            cupboard_orders[index][0] += order[0]
                            break
                    else:
                        cupboard_orders.append(order)
            else:
                print(f'INVALID order number {order_code} in row {row}')



def sorting_order_parts():
    sort_order = []
    
    #seperate items
    for parts in cupboard_orders:
        order = []
        order.append(parts[0])
        order.extend(parts[1].split('_'))
        sort_order.append(order)

    #sorted list
    sort_order = sorted(sort_order, key = lambda x: (x[1], x[5], x[4], x[2]))

    return sort_order
    

def write_excel_orderlist(items):
    wb = openpyxl.Workbook()
    wb.create_sheet("PARTS STYLE DETAILS")
    wb.create_sheet("PARTS COLOR IN STYLES")
    wb.create_sheet("PARTS COLOR DETAILS")
    ws = [ wb["PARTS STYLE DETAILS"], wb["PARTS COLOR IN STYLES"], wb["PARTS COLOR DETAILS"] ]

    write_parts_for_workshop(ws, items)
    adjust_column_width(ws)

    wb.save('output.xlsx')
    
        
    

        

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <excel_file>')
        exit(1)
    path = sys.argv[1]
    workbook = access_excel(path)
    extract_VANITY_INFO()
    get_Orders_MAIN_SHEET()
    items = sorting_order_parts()
    write_excel_orderlist(items)
    
    
   
