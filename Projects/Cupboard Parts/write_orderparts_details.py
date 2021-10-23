import sys
import openpyxl
from openpyxl.styles import *


def adjust_column_width(work_sheets):
    for ws in work_sheets:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if length > 20:
                length = 20
            ws.column_dimensions[column_cells[0].column_letter].width = length+5
            

def write_parts_for_workshop(work_sheets, items):
    '''
    Seperate all items by material 
    '''
    materials = [ x[1] for x in items ]
    materials = list(set(materials))
    materials.sort()

    start_row = 1
    for material in materials:
        items_for_material = [x for x in items if x[1] == material]
        start_row += write_material_items(work_sheets, start_row, items_for_material)


def write_material_items(work_sheets, start_row, items):
    '''
    Write material & style header and seperate items by part name
    '''

    row = start_row
    for ws in work_sheets:
        ws.cell(row, 1).value = "MATERIAL"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = items[0][1]
        ws.cell(row, 2).font = Font(bold=True)
    row += 1

    styles = [ x[2] for x in items ]
    styles = list(set(styles))
    styles.sort()

    colors = [ x[3] for x in items ]
    colors = list(set(colors))
    colors.sort()

    for ws in work_sheets:
        ws.cell(row, 1).value = "ITEM"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "SIZE"
        ws.cell(row, 2).font = Font(bold=True)

        if ws.title == "PARTS COLOR DETAILS":
            for i, color in enumerate(colors):
                ws.cell(row, i+3).value = color
                ws.cell(row, i+3).font = Font(bold=True)
        else:
            for i, style in enumerate(styles):
                ws.cell(row, i+3).value = style
                ws.cell(row, i+3).font = Font(bold=True)
    row += 1

    parts = [ x[5] for x in items ]
    parts = list(set(parts))
    parts.sort()
    part_names = []

    #set the order of part names
    for part in parts:
        if 'DOOR' in parts:
            part_names.insert(0,'DOOR')
            parts.remove('DOOR')
        if 'DRAWER' in parts:
            part_names.insert(1,'DRAWER')
            parts.remove('DRAWER')
        part_names.append(part)

    for part_name in part_names:
        part_items = [x for x in items if x[5] == part_name]
        row = write_part_names(work_sheets, row, part_items, styles,colors)
        row += 3

    return row

def write_part_names(work_sheets, start_row, items, styles, colors):
    '''
    Write detail for a specific part. Rows contain sizes and columns contain styles
    '''
    sizes = [ x[4] for x in items ]
    sizes = list(set(sizes))
    sizes = sorted(sizes, key = lambda x: int(x.split()[0]))

    #excel cell styles
    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))

    

    for ws in work_sheets:
        row = start_row
        ws.cell(row,1).value = items[0][5]
        
        #for total count
        tot_count_colors = [0 for x in range(len(colors))]
        tot_count_styles = [0 for x in range(len(styles))]
        
        for size in sizes:
            ws.cell(row,2).value = size
            size_items  = [x for x in items if x[4] == size]

            if ws.title == "PARTS COLOR DETAILS":
                for i,color in enumerate(colors):
                    color_items = [x for x in size_items if x[3] == color]
                    count = [x[0] for x in color_items]
                    if sum(count):
                        ws.cell(row,i+3).value = sum(count)
                    
                    #count of parts
                    tot_count_colors[i] += sum(count)
                    ws.cell(start_row+len(sizes),i+3).value = tot_count_colors[i]
                    ws.cell(start_row+len(sizes),i+3).border = border_style
                    
            else:
                for i,style in enumerate(styles):
                    style_items = [x for x in size_items if x[2] == style]
                    
                    if ws.title == "PARTS STYLE DETAILS":
                        count = [x[0] for x in style_items]
                        if sum(count):
                            ws.cell(row,i+3).value = sum(count)
                            
                        #count of parts
                        tot_count_styles[i] += sum(count)
                        ws.cell(start_row+len(sizes),i+3).value = tot_count_styles[i]
                        ws.cell(start_row+len(sizes),i+3).border = border_style 
                        
                    if ws.title == "PARTS COLOR IN STYLES":    
                        values = ''
                        for color in colors:
                            color_items = [x for x in style_items if x[3] == color]
                            count = [x[0] for x in color_items]
                            if sum(count):
                                values += str(sum(count)) + ' - ' + color + '\n'
                                ws.cell(row,i+3).value = values
                                if len(style_items) > 1:
                                    row_count = values.count('\n') + 1
                                    ws.cell(row,i+3).alignment = Alignment(wrap_text=True)
                                    #ws.row_dimensions[row].height = 15 * row_count
                                    
                            #count of parts
                            tot_count_styles[i] += sum(count)
                            ws.cell(start_row+len(sizes),i+3).value = tot_count_styles[i]
                            ws.cell(start_row+len(sizes),i+3).border = border_style
            row += 1

        #total of parts
        ws.cell(start_row+len(sizes),1).value = 'Total'
        ws.cell(start_row+len(sizes),2).border = border_style
        if ws.title == "PARTS COLOR DETAILS":
            ws.cell(start_row+len(sizes),2).value = sum(tot_count_colors)
        else:
            ws.cell(start_row+len(sizes),2).value = sum(tot_count_styles)
            
            
    return row



if __name__ == "__main__":
    pass
