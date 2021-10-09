from openpyxl import Workbook, load_workbook

# Initial measure set
cupboard_parts = {}
material = {}
cupboard_orderParts = {}
workbook = None
cupboord_parts = ['DOOR', 'DRAWER', 'CLASSIC KICK', 'MOULDING', 'TOWER MOULDING']

"""
material = {
    "MAPLE":{},
    "OAK":{},
    "MDF":{},
    "HARDROCK":{}
    }
    
door_style = {
    "RICHMOND FLAT":[],
    "RICHMOND RAISED":[],
    "ASHTON FLAT":[],
    "ASHTON RAISED":[],
    "MONACO FLAT":[],
    "MONACO RAISED":[],
    "HAMPTON FLAT":[],
    "HAMPTON RAISED":[],
    "SHAKER":[],
    "BEADED SHAKER":[],
    "FUSION":[],
    "URBAN":[],
    "MADISON":[],
    "SIERRA":[],
    "KENZO FLAT":[],
    "KENZO RAISED":[],
    "LOTUS FLAT":[],
    "LOTUS RAISED":[],
    "VISTA FLAT":[],
    "VISTA RAISED":[],
    "CAPRICE FLAT":[],
    "CAPRICE RAISED":[],
    "RUBY":[]
    }
"""

#set the data as same format and cleanup unnecessary characters
def cleanup_data(measure):
    if measure:
        if '-' in measure:
            measure = measure.replace('-','')
        if 'X' in measure:  
            measure = measure.replace('X',' x ')
        #set single space    
        measure_str = ' '.join(measure.split())
        measure_part = measure_str.split('[')
        #remove None values
        measure_part = list(filter(None, measure_part))
        measure_part = [x.replace(']', '').strip() for x in measure_part]
        
        return measure_part

    

#set the parts measures and count related to the code 
def set_cupboard_parts(code, measure):
    measure_list = []

    #seperate the count and measure      
    for n in range(len(measure)):
        measure_list.append([])
        if(measure[n][0].isdigit()):
            measure_list[n].append(int(measure[n][0]))
            measure_list[n].append(measure[n][1:].strip())
        else:
            measure_list[n].append(measure[n])

    #set measures with code              
    #measure_list = [tuple(x) for x in measure_list]
    cupboard_parts[code] = measure_list

    

#get excel access
def access_excel(path):
    xlsx_file = path
    wb = load_workbook(xlsx_file)

    return wb
    

#Access excel file to extract cupboard parts default measures data           
def extract_VANITY_INFO():
    ws = workbook["VANITY INFO"]

    for row in range(1, ws.max_row+1):
        code = ws.cell(row,3).value
        if code != None:
            if(isinstance(code,int)):
                measure = ws.cell(row,4).value
                measure_part = cleanup_data(measure)
                set_cupboard_parts(code, measure_part)

                
                

#get orders details
def get_Orders_MAIN_SHEET():
    ws = workbook["Main Sheet"]
    material = {}
    order_list = []
    
    for row in range(1, ws.max_row+1):
        order_code = ws.cell(row,5).value
        if order_code != None:
            if(isinstance(order_code,int)):
                order_list.append(order_code)
                material_name = ws.cell(row,8).value
                style = ws.cell(row,7).value 
                if material_name in material:
                    if style in material[material_name]:
                        material[material_name][style].append(order_code)
                    else:
                        material[material_name][style] = []
                        material[material_name][style].append(order_code)
                else:
                    material[material_name] = {}
                    material[material_name][style] = []
                    material[material_name][style].append(order_code)

    return material
                    

                
    

#count  the same measure parts
def cupboard_parts_count(order_list):
    cupboardParts_orderList = []

    for order in order_list:
        if order in cupboard_parts:
            for m_list in cupboard_parts[order]:
                check_list = cupboardParts_orderList.copy()
                if cupboardParts_orderList:
                    for o_list in cupboardParts_orderList:
                        if o_list[1] == m_list[1]:
                            index = cupboardParts_orderList.index(o_list)
                            #cupboardParts_orderList[index] = list(cupboardParts_orderList[index])
                            cupboardParts_orderList[index][0] += m_list[0]
                            #cupboardParts_orderList[index] = tuple(cupboardParts_orderList[index])
                            break
                        else:
                            check_list.remove(o_list)
                    else:
                        if not check_list:
                            cupboardParts_orderList.append(m_list)
                else:
                    cupboardParts_orderList.append(m_list)
        else:
            print('Wrong cupboard code !', order)
    
    
    return cupboardParts_orderList





#set the orders variations
def cupboard_order_variation():
    cupboard_orderParts = material.copy()
    
    for key in material:
        for style in material[key]:
            order = cupboard_parts_count(material[key][style])
            cupboard_orderParts[key][style] = {}
            for measure in order:
                for parts in cupboord_parts:
                    if parts in measure[1]:
                        m_list = measure.copy()
                        m_list[1] = m_list[1].replace(parts,'').strip()
                        if parts in cupboard_orderParts[key][style]:
                            cupboard_orderParts[key][style][parts].append(m_list)
                        else:
                            cupboard_orderParts[key][style][parts] = []
                            cupboard_orderParts[key][style][parts].append(m_list)       
    """
                   
    #Print values
    for key in cupboard_orderParts:
        print(key , ':')
        for style in cupboard_orderParts[key]:
            print(style , ':' )
            for parts in cupboard_orderParts[key][style]:
                print(parts,':', cupboard_orderParts[key][style][parts])
            print()
        print('===========================')
    """

    return cupboard_orderParts





#seperate cupboard parts according to style and material
def cupboardParts_order_style_count():
    order_count = {}

    #set order variation dictionary
    for key in cupboard_orderParts:
        order_count[key] = {}
        for style in cupboard_orderParts[key]:
            for parts in cupboard_orderParts[key][style]:
                order_count[key][parts] = {}

    #pass the values to related parts and style
    for key in order_count:
        for parts in order_count[key]:
            for style in cupboard_orderParts[key]:
                for o_parts in cupboard_orderParts[key][style]:
                    measure = cupboard_orderParts[key][style][o_parts]
                    if parts == o_parts:
                        if style in order_count[key][parts]:
                            order_count[key][parts][style].append(measure)
                        else:
                            order_count[key][parts][style] = []
                            order_count[key][parts][style].append(measure)
    
                   
    #Print values
    for key in order_count:
        print(key , ':')
        for parts in order_count[key]:
            print(parts , ':' )
            for style in order_count[key][parts]:
                print(style,':', order_count[key][parts][style])
            print()
        print('===========================')
    
    
        
    

if __name__ == "__main__":
    path = 'cupboard_parts.xlsx'
    workbook = access_excel(path)
    extract_VANITY_INFO()
    material = get_Orders_MAIN_SHEET()
    cupboard_orderParts = cupboard_order_variation()
    cupboardParts_order_style_count()
    
