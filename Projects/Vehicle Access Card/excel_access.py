import json
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

#file names
CONFIG_FILE_NAME = ".config.json"
OUTPUT_FILE_NAME = "Reminder.xlsx"

#Parameter Initialization
CONFIG_PARAMS = {
    "file_path"         : "",
    "email"             : "",
    "reminder_days"     : "",
    "expiry_title"      : "",
    "output_columns"    : []
}



    
#update json file
def update_configfile(config_params,file_name=CONFIG_FILE_NAME):
    with open(file_name, 'w') as fp:
        json.dump(config_params, fp)
        

#retrive parameter from json file
def retrive_configfile(file_name=CONFIG_FILE_NAME):
    with open(file_name) as fp:
        params = json.load(fp)
    return params


#Get Excel activated column number & start row number
def parse_xlsx_header(cfg_params):
    error_msg = None
    data_locations = {
        "work_sheet" : None,
        "data_start_row": None,
        "columns": None,
    }
    data_locations["columns"] = [None]
    data_locations["columns"].extend([None for x in cfg_params["output_columns"]])
    num_locations = len(data_locations["columns"])

    xlsx_file = cfg_params["file_path"]
    wb = load_workbook(xlsx_file)
    ws = wb.active
    data_locations["work_sheet"] = ws

    for row in range(1, ws.max_row+1):
        for column in range(1, ws.max_column+1):
            if (ws.cell(row, column).value == cfg_params["expiry_title"]):
                data_locations["columns"][0] = column
                data_locations["data_start_row"] = row + 1
                num_locations -= 1
            
            for column_label in cfg_params["output_columns"]:
                if ws.cell(row, column).value == column_label:
                    index = cfg_params["output_columns"].index(column_label)
                    data_locations["columns"][1+index] = column
                    num_locations -= 1
                
            if (num_locations == 0):
                break
        if (num_locations == 0):
            break

    #Show error when column is not exist
    if(None is data_locations["columns"][0]):
        error_msg = 'Expiry Title is Not Found !'
    elif(None in data_locations["columns"]):
        error_msg = 'Output Column is Not Found !'

    return (error_msg, data_locations)


#Get Excel Data sets related columns
def get_xlsx_data(xlsx_locations):
    xlsx_data = []
    ws = xlsx_locations["work_sheet"]

    for row in range(xlsx_locations["data_start_row"], ws.max_row+1):
        row_data = []
        for column in xlsx_locations["columns"]:
            value = ws.cell(row, column).value
            if(isinstance(value,datetime)):
                value = value.date()
            row_data.append(value)
        if row_data[0] != None:
            xlsx_data.append(row_data)

    return xlsx_data



#Filter data regarding parameter
def filter_data(xlsx_data,cfg_params):
    num_expiry_days = cfg_params["reminder_days"]
    filtered_data = []
    #reminding days before expired
    remind_day = datetime.now().date() - timedelta(days = num_expiry_days)
    
    for data in xlsx_data:
        if(data[0] == remind_day):
            filtered_data.append(data)       
                   
    return filtered_data


#Write filtered data into output excel file
def write_xlsx_file(xlsx_data, cfg_params):
    output_file = OUTPUT_FILE_NAME
    
    wb = Workbook()
    ws = wb.active
    
    #write headers
    header = []
    header.append(cfg_params["expiry_title"])
    header.extend(cfg_params["output_columns"])
    ws.append(header)

    for row in xlsx_data:
       ws.append(row)

    wb.save(output_file)

    return output_file
   
 
