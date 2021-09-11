import json
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date
from gui import main_func,getdata


CONFIG_FILE_NAME = ".config.json"

# Get following parameters from GUI inputs.
"""
config_params = {
    "receiver_email" : "ravitharan@gmail.com",
    "notification_days" : 5,
    "expiry_title" : "Valid till",
    "output_columns" : ["Name", "Access card no", "Vehicle no"],
}
"""


def getConfigData(data): 
    para = ["expiry_title", "notification_days", "output_columns",  "receiver_email"]
    config_params = {}
    
    for key in para:
        for val in data:
            if(key == 'output_columns'):
                config_params[key] = list(val.split(','))
            else:
                config_params[key] = val
            data.remove(val)
            break
        
    update_configfile(CONFIG_FILE_NAME,config_params)
    
    
    

    

def update_configfile(file_name, config_params):
    with open(file_name, 'w') as fp:
        json.dump(config_params, fp)
        print('update_configfile')

def retrive_configfile(file_name):
    with open(file_name) as fp:
        params = json.load(fp)
        print('retrive_configfile')
    return params

def parse_xlsx_header(xlsx_file, cfg_params):
    '''
    Parse xlsx file and returen column number for expiry date and other output
    columns. Also returning data start row number
    '''
    data_locations = {
        "work_sheet" : None,
        "data_start_row": None,
        "columns": None,
    }
    data_locations["columns"] = [None]
    data_locations["columns"].extend([None for x in cfg_params["output_columns"]])
    num_locations = len(data_locations["columns"])

    wb = load_workbook(xlsx_file)
    ws = wb.active
    data_locations["work_sheet"] = ws

    for row in range(1, ws.max_row+1):
        for column in range(1, ws.max_column+1):
            if (ws.cell(row, column).value == cfg_params["expiry_title"]):
                data_locations["columns"][0] = column
                data_locations["data_start_row"] = row + 1
                num_locations -= 1
            for column_label in cfg_params["output_columns"]:
                if ws.cell(row, column).value == column_label:
                    index = cfg_params["output_columns"].index(column_label)
                    data_locations["columns"][1+index] = column
                    num_locations -= 1
            if (num_locations == 0):
                break
        if (num_locations == 0):
            break

    return data_locations


def get_xlsx_data(xlsx_locations):
    '''
    Return expiry date and other output columns in a list from the excel file
    '''
    xlsx_data = []
    ws = xlsx_locations["work_sheet"]

    for row in range(xlsx_locations["data_start_row"], ws.max_row+1):
        row_data = []
        for column in xlsx_locations["columns"]:
            value = ws.cell(row, column).value
            if(isinstance(value,datetime)):
                value = value.date()
            row_data.append(value)
        if row_data[0] != None:
            xlsx_data.append(row_data)

    return xlsx_data



def filter_data(xlsx_data,cfg_params):
    '''
    Filter xlsx_data within num_expiry_days. Sort the filtered data
    '''
    num_expiry_days = int(cfg_params["notification_days"])
    filtered_data = []
    #reminding days before expired
    remind_day = datetime.now().date() - timedelta(days = num_expiry_days)
    
    for data in xlsx_data:
        if(data[0] < remind_day):
            filtered_data.append(data)       
                   
    return filtered_data

    
def write_xlsx_file(xlsx_data, cfg_params):
    '''
    Write filtered data into output xlsx file
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = "Reminding Clients"
    
    #write headers
    header = []
    header.append(cfg_params["expiry_title"])
    header.extend(cfg_params["output_columns"])
    ws.append(header)

    for row in xlsx_data:
       ws.append(row)

    wb.save('Reminder_VehicleCard_Data.xlsx')
   
 
if __name__ == '__main__':
    main_func()
    data = getdata()
    print(data)
    if(data):
        getConfigData(data)
        
    config_params = retrive_configfile(CONFIG_FILE_NAME)
    if(config_params):
        locations = parse_xlsx_header('DMM Access cards details.xlsx', config_params)
        data = get_xlsx_data(locations)
        filter_data = filter_data(data, config_params)
        print(filter_data)
        write_xlsx_file(filter_data,config_params)
   
