import json
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

#file names
CONFIG_FILE_NAME = ".config.json"
OUTPUT_FILE_NAME = "Reminder.xlsx"

#Parameter Initialization
CONFIG_PARAMS = {
    "file_path"         : "",
    "email"             : "",
    "reminder_days"     : "",
    "expiry_title"      : "",
    "output_columns"    : []
}



    
#update json file
def update_configfile(config_params,file_name=CONFIG_FILE_NAME):
    with open(file_name, 'w') as fp:
        json.dump(config_params, fp)
        

#retrive parameter from json file
def retrive_configfile(file_name=CONFIG_FILE_NAME):
    with open(file_name) as fp:
        params = json.load(fp)
    return params


#Get Excel activated column number & start row number
def parse_xlsx_header(file_path):

    #xlsx_file = cfg_params["file_path"]
    xlsx_file = file_path
    wb = load_workbook(xlsx_file)
    ws = wb.active
    worksheet = ws
    col_names = []
    
    for col in range(1, ws.max_column+1):
        col_names.append(ws.cell(1,col).value)
        #print(ws.cell(1,col).value)
    

    return (col_names, worksheet)


#Get Excel Data sets related columns
def get_xlsx_data(cfg_params,worksheet):
    xlsx_data = []
    ws = worksheet
    col_num = []
    error_msg = None


    for col in range(1, ws.max_column+1):
        if (ws.cell(1,col).value == cfg_params["expiry_title"]):
            col_num.append(col)
            
    for col in range(1, ws.max_column+1):
        if ws.cell(1,col).value in cfg_params["output_columns"]:
            if(col != col_num[0]):
                col_num.append(col)
            
    
    for row in range(2, ws.max_row+1):
        row_data = []
        for column in col_num:
            value = ws.cell(row, column).value
            if(isinstance(value,datetime)):
                value = value.date()
            row_data.append(value)
        if row_data[0] != None:
            xlsx_data.append(row_data)
        
            

    return (xlsx_data)



#Filter data regarding parameter
def filter_data(xlsx_data,cfg_params):
    num_expiry_days = cfg_params["reminder_days"]
    filtered_data = []
    error_msg = None
    #reminding days before expired
    remind_day = datetime.now().date() - timedelta(days = num_expiry_days)
    
    for data in xlsx_data:
        
        #Check Expiry title data is DATE type
        if(isinstance(data[0],date)):
            if(data[0] < remind_day):
                filtered_data.append(data)
        else:
            error_msg = 'Expiry Title is not DATE type !'
            break
                   
    return (error_msg,filtered_data)


#Write filtered data into output excel file
def write_xlsx_file(xlsx_data, cfg_params):
    output_file = OUTPUT_FILE_NAME
    
    wb = Workbook()
    ws = wb.active
    
    #write headers
    header = []
    header.append(cfg_params["expiry_title"])
    header.extend(cfg_params["output_columns"])
    ws.append(header)

    for row in xlsx_data:
       ws.append(row)

    wb.save(output_file)

    return output_file
   
 
